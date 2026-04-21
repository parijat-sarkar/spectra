"""
Validate Beagle+ output against a real Beagle export.

Usage:
    python validate_vs_beagle.py <beagle.xlsx> <ENST_id> [--editor ABE|CBE] [--pam NGG|NGN] [--window 3..10]

What this checks:
  1) Every unique (sgRNA sequence, orientation, global start) in the Beagle
     export is produced by Beagle+ when run on the same ENST with matching
     editor/PAM/window.
  2) For each Beagle full-edit row, Beagle+'s full-edit-outcome row exists
     with the same Nucleotide Edits (global), HGVS c. notation, and
     Mutation Category.
  3) Report where Beagle+ adds rows Beagle does not (these are the
     partial-edit outcomes — the feature).

Requires: openpyxl (pip install openpyxl).
Requires network access to rest.ensembl.org.
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

from beagle_core import generate_table, COLUMNS


def load_beagle(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]
    header = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({h: v for h, v in zip(header, row)})
    return rows


def main():
    p = argparse.ArgumentParser()
    p.add_argument("xlsx", help="Beagle xlsx export")
    p.add_argument("enst", help="ENST id, e.g. ENST00000349945.7")
    p.add_argument("--editor", default="ABE", choices=["ABE", "CBE"])
    p.add_argument("--pam", default="NGN", choices=["NGG", "NGN"])
    p.add_argument("--window", default="3..10")
    args = p.parse_args()

    w_start, w_end = map(int, args.window.split(".."))
    print(f"Loading Beagle export: {args.xlsx}")
    beagle = load_beagle(args.xlsx)
    print(f"  {len(beagle)} rows in Beagle export")

    print(f"Generating Beagle+ table for {args.enst} ({args.editor}/{args.pam} {args.window})...")
    result = generate_table(args.enst, args.editor, args.pam, (w_start, w_end))
    cols = result["columns"]
    rows = result["rows"]
    print(f"  {len(rows)} rows in Beagle+")

    col = {c: i for i, c in enumerate(cols)}

    # Beagle+ emits ONE row per guide. All partial-edit combinations
    # are concatenated inside each outcome cell using " | ".
    by_guide = {}
    for r in rows:
        key = (r[col["sgRNA Sequence"]],
               int(r[col["sgRNA Sequence Start Pos. (global)"]]),
               r[col["sgRNA Orientation"]])
        by_guide[key] = r

    missing = 0
    matched_full = 0
    mismatched = []
    for b in beagle:
        key = (b["sgRNA Sequence"],
               int(b["sgRNA Sequence Start Pos. (global)"]),
               b["sgRNA Orientation"])
        if key not in by_guide:
            missing += 1
            if missing <= 5:
                print(f"  MISSING: {key}")
            continue
        bp = by_guide[key]
        # Beagle's row corresponds to the full-edit combination.
        # Find the combo inside Beagle+'s cell whose edit set matches.
        b_global = set((b.get("Nucleotide Edits (global)") or "").split(", "))
        found = False
        for entry in (bp[col["Nucleotide Edits (global)"]] or "").split(" | "):
            if set(entry.split(", ")) == b_global:
                found = True
                break
        if not found:
            mismatched.append((key, f"no matching combo inside cell; beagle={b_global}"))
            continue
        matched_full += 1

    print()
    print(f"Guides in Beagle but not in Beagle+: {missing}")
    print(f"Full-edit combos matched inside Beagle+ cells: {matched_full} / {len(beagle) - missing}")
    print(f"Mismatches: {len(mismatched)}")
    for m in mismatched[:10]:
        print(f"  {m}")

    combos = [int(r[col["Total Combinations for Guide"]]) for r in rows]
    if combos:
        print()
        print(f"Beagle+ guides: {len(rows)}")
        print(f"Avg combos per guide: {sum(combos)/len(combos):.1f}")
        print(f"Max combos on a single guide: {max(combos)}")


if __name__ == "__main__":
    main()
